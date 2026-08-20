from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import shutil
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "Datum",
    "Naam",
    "Onderwerp",
    "Medewerker",
    "Groepering",
    "Status",
    "Vlgd. contact",
    "Contact",
    "Gsm",
    "Memo",
]
DATE_PATTERN = re.compile(r"(?<!\d)(\d{2}/\d{2}/(?:\d{2}|\d{4}))(?!\d)")

STATUS_MAP = {
    "open": "Lopend",
    "lopend": "Lopend",
    "actief": "Lopend",
    "in behandeling": "Lopend",
    "wachtend": "Wachtend",
    "waiting": "Wachtend",
    "afgesloten": "Afgesloten",
    "afgewerkt": "Afgesloten",
    "completed": "Afgesloten",
    "closed": "Afgesloten",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _normalize_identity(value: str) -> str:
    return " ".join(value.casefold().split())


def _identity_key(customer: str, subject: str) -> str:
    return f"{_normalize_identity(customer)}|{_normalize_identity(subject)}"


def _date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    value_text = _text(value)
    if not value_text:
        return ""
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value_text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _memo_date(value: str) -> str | None:
    try:
        parsed = datetime.strptime(value, "%d/%m/%Y")
    except ValueError:
        try:
            parsed = datetime.strptime(value, "%d/%m/%y")
        except ValueError:
            return None
    if parsed.year < 2000:
        parsed = parsed.replace(year=parsed.year + 100)
    return parsed.date().isoformat()


def _status(value: str) -> tuple[str, bool]:
    normalized = " ".join(value.casefold().split())
    mapped = STATUS_MAP.get(normalized)
    return mapped or "Lopend", mapped is not None


def _parse_memo(memo: str, fallback_date: str, event_type: str) -> tuple[list[dict[str, Any]], bool, bool]:
    matches = list(DATE_PATTERN.finditer(memo))
    dated_entries: list[dict[str, Any]] = []
    invalid_date_found = False

    for match in matches:
        parsed_date = _memo_date(match.group(1))
        if parsed_date is None:
            invalid_date_found = True
            continue
        dated_entries.append({"position": match.start(), "date": parsed_date})

    if not dated_entries:
        return [
            {
                "event_date": fallback_date,
                "event_type": event_type or "Notitie",
                "notes": memo,
                "follow_up_date": None,
                "origin": "Excel Datum fallback",
            }
        ], True, invalid_date_found

    events: list[dict[str, Any]] = []
    prefix = memo[: dated_entries[0]["position"]].strip()
    for index, entry in enumerate(dated_entries):
        start = entry["position"] + len(next(match.group(1) for match in matches if match.start() == entry["position"]))
        end = dated_entries[index + 1]["position"] if index + 1 < len(dated_entries) else len(memo)
        notes = memo[start:end].strip()
        if index == 0 and prefix:
            notes = f"{prefix}\n{notes}".strip()
        events.append(
            {
                "event_date": entry["date"],
                "event_type": event_type or "Notitie",
                "notes": notes,
                "follow_up_date": None,
                "origin": "Memo date",
            }
        )
    return events, False, invalid_date_found


def _row_hash(row: dict[str, str]) -> str:
    payload = "\x1f".join(row.get(column, "") for column in EXPECTED_COLUMNS if column != "Medewerker")
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _manual_identity_keys(db_path: Path) -> set[str]:
    if not db_path.exists():
        return set()
    with sqlite3.connect(f"file:{db_path.resolve()}?mode=ro", uri=True) as conn:
        return {
            _identity_key(row[0], row[1])
            for row in conn.execute("SELECT customer, subject FROM dossiers WHERE source IS NULL OR source != 'Adsolut Communicator'")
        }


def build_report(input_path: Path, db_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(rows, ())]
    missing_columns = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    positions = {header: index for index, header in enumerate(headers)}
    proposed_dossiers: dict[str, dict[str, Any]] = {}
    row_reports: list[dict[str, Any]] = []
    hash_counts: Counter[str] = Counter()
    dated_memo_rows = 0
    fallback_rows = 0
    ambiguous_rows = 0
    missing_required_rows = 0
    status_mappings: list[dict[str, Any]] = []

    for row_number, values in enumerate(rows, start=2):
        row = {column: _text(values[positions[column]] if positions[column] < len(values) else "") for column in EXPECTED_COLUMNS}
        customer = row["Naam"]
        subject = row["Onderwerp"] or "leeg"
        row["Onderwerp"] = subject
        missing = [field for field, value in (("Naam", customer),) if not value]
        if missing:
            missing_required_rows += 1

        status, known_status = _status(row["Status"])
        status_mappings.append({"row": row_number, "source": row["Status"], "target": status, "known": known_status})
        row_hash = _row_hash(row)
        hash_counts[row_hash] += 1
        fallback_date = _date_value(row["Datum"])
        follow_up_date = _date_value(row["Vlgd. contact"])
        events, used_fallback, ambiguous = _parse_memo(row["Memo"], fallback_date, row["Groepering"])
        dated_memo_rows += not used_fallback
        fallback_rows += used_fallback
        ambiguous_rows += ambiguous

        identity = _identity_key(customer, subject)
        dossier_key = f"{identity}::row-{row_number}"
        dossier = proposed_dossiers.setdefault(
            dossier_key,
            {
                "identity_key": identity,
                "dossier_key": dossier_key,
                "source": "Adsolut Communicator",
                "customer": customer,
                "subject": subject,
                "contact": row["Contact"],
                "status": status,
                "follow_up_date": follow_up_date,
                "latest_row": row_number,
                "events": [],
                "source_rows": [],
            },
        )
        if row_number >= dossier["latest_row"]:
            dossier.update({"contact": row["Contact"], "status": status, "follow_up_date": follow_up_date, "latest_row": row_number})
        dossier["source_rows"].append(row_number)
        dossier["events"].extend(events)
        row_reports.append({"row": row_number, "identity_key": identity, "dossier_key": dossier_key, "row_hash": row_hash, "events": events, "ambiguous_memo": ambiguous, "missing_required_fields": missing})

    manual_keys = _manual_identity_keys(db_path)
    for dossier in proposed_dossiers.values():
        dossier["warnings"] = []
        if dossier["identity_key"] in manual_keys:
            dossier["warnings"].append("Possible match with an existing manual dossier; no automatic merge proposed.")

    return {
        "dry_run": True,
        "input_file": str(input_path),
        "total_excel_rows": len(row_reports),
        "proposed_dossier_count": len(proposed_dossiers),
        "proposed_event_count": sum(len(row["events"]) for row in row_reports),
        "rows_with_dated_memo_entries": dated_memo_rows,
        "rows_using_excel_datum_fallback": fallback_rows,
        "ambiguous_memo_rows": ambiguous_rows,
        "duplicate_row_hashes": sorted(hash_value for hash_value, count in hash_counts.items() if count > 1),
        "status_mappings": status_mappings,
        "rows_with_missing_required_fields": missing_required_rows,
        "ignored_columns": ["Medewerker", "Gsm"],
        "proposed_dossiers": list(proposed_dossiers.values()),
        "rows": row_reports,
    }


def _validate_commit_report(report: dict[str, Any]) -> None:
    if report["ambiguous_memo_rows"]:
        raise ValueError("Commit refused: ambiguous memo rows must be resolved first.")
    if report["rows_with_missing_required_fields"]:
        raise ValueError("Commit refused: rows with missing required fields must be resolved first.")
    if any(mapping["target"] not in {"Lopend", "Wachtend", "Afgesloten"} for mapping in report["status_mappings"]):
        raise ValueError("Commit refused: at least one status cannot be mapped safely.")
    for row in report["rows"]:
        for event in row["events"]:
            if not event["event_date"]:
                raise ValueError(f"Commit refused: row {row['row']} has no valid event date.")


def _backup_database(db_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = db_path.parents[1] / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{db_path.stem}_{timestamp}{db_path.suffix}"
    shutil.copy2(db_path, backup_path)
    return backup_path


def commit_report(report: dict[str, Any], db_path: Path) -> dict[str, Any]:
    _validate_commit_report(report)
    if not db_path.exists():
        raise ValueError(f"Database does not exist: {db_path}")

    backup_path = _backup_database(db_path)
    dossiers_created = 0
    events_created = 0
    duplicate_rows_skipped = 0

    try:
        with sqlite3.connect(db_path) as conn:
            conn.execute("BEGIN")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS communicator_import_rows (
                    source_row_hash TEXT PRIMARY KEY,
                    source TEXT NOT NULL,
                    source_file TEXT NOT NULL,
                    source_row_number INTEGER NOT NULL,
                    dossier_id TEXT NOT NULL,
                    event_ids TEXT NOT NULL,
                    imported_at TEXT NOT NULL
                )
                """
            )

            for row in report["rows"]:
                existing = conn.execute(
                    "SELECT 1 FROM communicator_import_rows WHERE source_row_hash = ?",
                    (row["row_hash"],),
                ).fetchone()
                if existing is not None:
                    duplicate_rows_skipped += 1
                    continue

                dossier = next(item for item in report["proposed_dossiers"] if item["dossier_key"] == row["dossier_key"])
                dossier_id = hashlib.sha256(f"{report['input_file']}\x1f{row['row'] }".encode("utf-8")).hexdigest()
                now = datetime.utcnow().isoformat()
                conn.execute(
                    """
                    INSERT INTO dossiers (
                        id, customer, contact, subject, status, follow_up_date,
                        source, external_id, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        dossier_id,
                        dossier["customer"],
                        dossier["contact"],
                        dossier["subject"],
                        dossier["status"],
                        dossier["follow_up_date"],
                        "Adsolut Communicator",
                        None,
                        now,
                        now,
                    ),
                )
                dossiers_created += 1

                event_ids: list[str] = []
                for event_index, event in enumerate(row["events"]):
                    event_id = hashlib.sha256(f"{row['row_hash']}\x1f{event_index}".encode("utf-8")).hexdigest()
                    conn.execute(
                        """
                        INSERT INTO dossier_events (
                            id, dossier_id, event_date, event_type, notes,
                            follow_up_date, status_change, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            event_id,
                            dossier_id,
                            event["event_date"],
                            event["event_type"],
                            event["notes"],
                            None,
                            None,
                            now,
                        ),
                    )
                    event_ids.append(event_id)
                    events_created += 1

                conn.execute(
                    """
                    INSERT INTO communicator_import_rows (
                        source_row_hash, source, source_file, source_row_number,
                        dossier_id, event_ids, imported_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row["row_hash"],
                        "Adsolut Communicator",
                        report["input_file"],
                        row["row"],
                        dossier_id,
                        json.dumps(event_ids),
                        now,
                    ),
                )
    except Exception as exc:
        raise RuntimeError(
            json.dumps(
                {
                    "dry_run": False,
                    "committed": False,
                    "backup_database_path": str(backup_path),
                    "rows_processed": 0,
                    "dossiers_created": 0,
                    "events_created": 0,
                    "duplicate_source_rows_skipped": duplicate_rows_skipped,
                    "failures": [str(exc)],
                    "rollback_status": "rolled back",
                },
                ensure_ascii=False,
            )
        ) from exc

    return {
        "dry_run": False,
        "committed": True,
        "backup_database_path": str(backup_path),
        "rows_processed": len(report["rows"]),
        "dossiers_created": dossiers_created,
        "events_created": events_created,
        "duplicate_source_rows_skipped": duplicate_rows_skipped,
        "failures": [],
        "rollback_status": "not needed",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Adsolut Communicator importer; dry-run by default, database writes require --commit.")
    parser.add_argument("input", type=Path, help="Path to the Communicator .xlsx export")
    parser.add_argument("--database", type=Path, default=Path(__file__).resolve().parents[3] / "database" / "actions.db")
    parser.add_argument("--commit", action="store_true", help="Commit the validated import; omitted by default")
    args = parser.parse_args()
    report = build_report(args.input, args.database)
    if not args.commit:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return
    print(json.dumps(commit_report(report, args.database), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
